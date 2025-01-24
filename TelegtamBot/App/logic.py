import openpyxl

def Percentage_of_Dz_Verified(number_verified, number_plann):
    number_average_score: float = (number_verified /number_plann)*100

    number_average_score_lod = (number_average_score * 10) % 10

    if int(number_average_score_lod) >= 5.0:

        return int(number_average_score + 1)

    else:

        return int(number_average_score)

def search_verified_homework(file,search_text_data):
        list_text = []
        list_search =["Проверено","План"]
        worbook = openpyxl.open("Отчет по домашним заданиям.xlsx")
        worksheet = worbook.active
        col_data = 0
        col_teacher= 0
        col_arguments = []

        for row in range(1, worksheet.max_row):
            for col in range(0,worksheet.max_column):
               if worksheet[row][col].value == search_text_data:

                    col_data = col

                    print(col_data)

               elif worksheet[row][col].value == "ФИО преподавателя":

                   col_teacher = col

                   print(col_teacher)

        for row in range(1, worksheet.max_row):
            for col in range(col_data, worksheet.max_column):
                if (worksheet[row][col].value == list_search[0]
                    or worksheet[row][col].value == list_search[1]) :

                    col_arguments.append(col)
                    print(col_arguments)

                elif(len(col_arguments)==2):
                    break

        for row in range(3, worksheet.max_row):
            if col_data == 0 or col_teacher == 0 and len(col_arguments) == 0:
                text = "Таких данных нет  в документе "
                list_text.append(text)
                return list_text
            else:
                worksheet_row_col_fio_teacher = worksheet[row][col_teacher].value
                worksheet_row_col_relet_1 = int(worksheet[row][col_arguments[0]].value)
                worksheet_row_col_relet_2 = int(worksheet[row][col_arguments[1]].value)
                if Percentage_of_Dz_Verified(worksheet_row_col_relet_1, worksheet_row_col_relet_2) < 75:
                    text = f"""Добрый день - {worksheet_row_col_fio_teacher}. У вас не выполнена норма по проверке ДЗ студентов. Нужно исправить это.У вас такой процент проверки ДЗ: {Percentage_of_Dz_Verified(worksheet_row_col_relet_1, worksheet_row_col_relet_2)}%."""
                    list_text.append(text)
                else:
                    text = "ок"
        worbook.close()
        return list_text

def search_issued_homework(file,search_text_data):
        list_text = []
        list_search =["Выдано","План"]
        worbook = openpyxl.open("Отчет по домашним заданиям.xlsx")
        worksheet = worbook.active
        col_data = 0
        col_teacher= 0
        col_arguments = []
        for row in range(1, worksheet.max_row):
            for col in range(0,worksheet.max_column):
               if worksheet[row][col].value == search_text_data:

                    col_data = col

                    print(col_data)
               elif worksheet[row][col].value == "ФИО преподавателя":

                   col_teacher = col

                   print(col_teacher)

        for row in range(1, worksheet.max_row):
            for col in range(col_data, worksheet.max_column):

                if (worksheet[row][col].value == list_search[0]
                    or worksheet[row][col].value == list_search[1]) :

                    col_arguments.append( col)
                    print(col_arguments)

                elif(len(col_arguments)==2):
                    break

        for row in range(3, worksheet.max_row):
            if col_data == 0 or col_teacher == 0 and len(col_arguments) == 0:
                text = "Таких данных нет  в документе "
                list_text.append(text)
                return list_text
            else:
                worksheet_row_col_fio_teacher = worksheet[row][col_teacher].value
                worksheet_row_col_homework_issued = int(worksheet[row][col_arguments[0]].value)
                worksheet_row_col_plan = int(worksheet[row][col_arguments[1]].value)
                if (Percentage_of_Dz_Verified(worksheet_row_col_homework_issued,
                                              worksheet_row_col_plan) < 70):
                    text = f"""Добрый день - {worksheet_row_col_fio_teacher}. У вас не выполнена норма по выдачи ДЗ студентов.Нужно исправить это.У вас такой процент выданных ДЗ: {Percentage_of_Dz_Verified(worksheet_row_col_homework_issued, worksheet_row_col_plan)}%."""
                    list_text.append(text)
                else:
                    text = "ок"

        worbook.close()
        return list_text

def search_average_rating_grop(file ):
    text = ""
    list_text = []
    list_procent= []
    list_teacher= []

    worbook = openpyxl.open("Отчетпопосещаемостистудентов.xlsx")
    worksheet = worbook.active

    col_percent = 0
    col_teacher= 0
    for row in range(1, worksheet.max_row):
        for col in range(0, worksheet.max_column):

            if worksheet[row][col].value == "Средняя посещаемость":
                col_percent = col
                print(col_percent)
                break

            elif worksheet[row][col].value == "ФИО преподавателя":
                col_teacher = col
                print(col_teacher)

    for row in range(3,worksheet.max_row+1):
        list_procent.append(worksheet[row][col_percent].value)
        list_teacher.append(worksheet[row][col_teacher].value)

    for len_list in range(len(list_procent)-1):
        if col_percent != 0 and col_teacher == 0:
            worksheet_row_col_fio_teacher = list_teacher[len_list]
            worksheet_row_col_procent_teacher = int(list_procent[len_list].strip("%"))
            worksheet_row_col_procent_teacher_finall = int(list_procent[len(list_procent) - 1])
            if worksheet_row_col_procent_teacher < worksheet_row_col_procent_teacher_finall:
                text = (f"""Добрый день {worksheet_row_col_fio_teacher}!У вас плохая посещаемость предмета на вашей паре.Вот такая {worksheet_row_col_procent_teacher}%.""")
                list_text.append(text)
            else:
                text = "Ок"
        else:
            text = "Таких данных нет  в документе "
            list_text.append(text)
            return list_text

    worbook.close()
    print(list_text)
    return list_text

def Average_score(number_homework,number_classwork ):
        number_average_score:float = (number_homework+number_classwork)/2
        number_average_score_lod = (number_average_score * 10)%10
        if int(number_average_score_lod) >= 5.0:
            return int(number_average_score+1)
        else:
            return int(number_average_score)

def search_student_assessment(file):
    list_text = []
    list_search = ["FIO","Группа","Classroom","Homework","Average score","Exam"]
    col_fio = 0
    col_group = 0
    col_assessment = []
    worbook = openpyxl.open("Отчетпостудентам.xlsx")
    worksheet = worbook.active
    for row in range(1, worksheet.max_row):
        for col in range(0, worksheet.max_column):

            if worksheet[row][col].value == list_search[0]:
                col_fio = col
                print(col_fio)
            elif worksheet[row][col].value == list_search[1]:
                col_group = col
                print(col_group)
            elif ((worksheet[row][col].value == list_search[2])or
                  (worksheet[row][col].value == list_search[3])or
                  (worksheet[row][col].value == list_search[4])):
                col_assessment.append(col)
                print(col_assessment)

    for row in range(2, worksheet.max_row):
        if col_fio == 0 and col_group != 0  and  len(col_assessment) != 0:
            worksheet_row_col_name_student = worksheet[row][col_fio].value
            worksheet_row_col_grop = worksheet[row][col_group].value
            worksheet_row_col_homework = int(worksheet[row][col_assessment[0]].value)
            worksheet_row_col_classroom = int(worksheet[row][col_assessment[1]].value)
            if (worksheet_row_col_homework < 3 or worksheet_row_col_classroom < 3
                    or Average_score(worksheet_row_col_homework, worksheet_row_col_classroom) < 3):
                text = f"""Cтудент:{worksheet_row_col_name_student} - {worksheet_row_col_grop} Дз- {worksheet_row_col_homework} или КЛ_Р- {worksheet_row_col_classroom}.Как поступить в данной ситуации?"""
                list_text.append(text)
            else:
                continue
        else:
            text = "Таких данных нет  в документе "
            list_text.append(text)
            return list_text

    worbook.close()
    return list_text

def search_percentage_of_homework_per_month(file):
    list_text = []
    list_search = ["FIO","Группа","Percentage Homework"]
    col_assessment = []
    worbook = openpyxl.open("Отчет по дз (6 задание) .xlsx")
    worksheet = worbook.active
    for row in range(1, worksheet.max_row):
        for col in range(0, worksheet.max_column):
            if worksheet[row][col].value == list_search[0]:
                col_assessment.append(col)
            elif worksheet[row][col].value == list_search[1]:
                col_assessment.append(col)
            elif worksheet[row][col].value == list_search[2]:
                col_assessment.append(col)
                break
    print(col_assessment)

    for row in range(2, worksheet.max_row):
        if len(col_assessment) != 0:
            worksheet_row_col_name_student = worksheet[row][col_assessment[0]].value
            worksheet_row_col_grop = worksheet[row][col_assessment[1]].value
            worksheet_row_col_homework = int(worksheet[row][col_assessment[2]].value)
            print(f"{worksheet_row_col_homework}")
            if worksheet_row_col_homework < 50:
                text = (
                f"""Студент:{worksheet_row_col_name_student}-{worksheet_row_col_grop} и процент дз:{worksheet_row_col_homework}%.""")
                list_text.append(text)
        else:
            text = "Таких данных нет  в документе "
            list_text.append(text)
            return list_text

    print(len(list_text))
    worbook.close()
    return list_text